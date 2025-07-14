import re
import pandas as pd
from datetime import datetime, time
from openpyxl import load_workbook
from copy import copy
from pathlib import Path
import shutil
import os
import json
from InquirerPy import inquirer
from InquirerPy.separator import Separator

# ========== CONFIGURACIÓN ==========

CONFIG_FILE = "exportador_inscritos.config.json"

IDIOMAS_VALIDOS = ["INGLÉS", "PORTUGUÉS", "ITALIANO", "QUECHUA"]

IDIOMA_ABBR = {
    "INGLÉS": "ING",
    "PORTUGUÉS": "PORT",
    "ITALIANO": "ITA",
    "QUECHUA": "QUE"
}
NIVEL_ABBR = {
    "Básico": "B",
    "Intermedio": "I",
    "Avanzado": "A"
}
MODALIDAD_ABBR = {
    "regular": "REG",
    "intensivo": "INT",
    "súperintensivo": "SINT",
    "superintensivo": "SINT",
    "repaso": "REP"
}
DIA_COD = {0: "L", 1: "M", 2: "X", 3: "J", 4: "V", 5: "S", 6: "D"}

COLUMNAS_FINALES = [
    "CODIGO", "Nivel", "Ciclo", "MODALIDAD", "DOCENTE", "IDIOMA", "DÍAS DETECTADOS",
    "HORARIO DETALLADO", "F. Inicio", "F. Fin", "Parcial", "Final", "Subida de notas",
    "N° Inscritos", "N° Esperado", "N° Aprobados", "N° Desaprobados",
    "N° No asistio (tiene 0)", "Destalle del curso"
]


def cargar_config():
    if Path(CONFIG_FILE).exists():
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def guardar_config(config):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=4)

def limpiar_consola():
    os.system('cls' if os.name == 'nt' else 'clear')

def pausar():
    input("\nPresiona ENTER para continuar...")

# ========== PARTE DE LIMPIEZA Y TRANSFORMACIÓN ==========

def clean_df_mes_idioma(excel_path, mes):
    # Lee todos los cursos (de todos los idiomas) del mes seleccionado
    df = pd.read_excel(excel_path, sheet_name=mes, skiprows=1)
    matricula_idx = df[df.iloc[:, 0].astype(str).str.upper().str.contains("MATRÍCULA")].index
    if not matricula_idx.empty:
        df = df.loc[:matricula_idx[0] - 1]

    df["IDIOMA"] = None
    idioma_actual = "INGLÉS"
    for i, row in df.iterrows():
        cod_val = str(row.get("CODIGO", "")).strip().upper()
        ciclo_val = str(row.get("CICLO", "")).strip().upper()
        for idioma in IDIOMAS_VALIDOS[1:]:
            if idioma in cod_val or idioma in ciclo_val:
                idioma_actual = idioma
                break
        df.at[i, "IDIOMA"] = idioma_actual
    df["IDIOMA"] = df["IDIOMA"].ffill()
    df = df[df["CODIGO"].notna()]
    df = df[~df["CODIGO"].astype(str).str.upper().isin(IDIOMAS_VALIDOS)]
    df = df[~df["CODIGO"].astype(str).str.upper().str.contains("CODIGO")]
    df["DOCENTE"] = df["DOCENTE"].ffill()

    # Nivel y Ciclo
    def extraer_nivel_y_ciclo(valor):
        valor = str(valor).strip().upper()
        if "REPASO" in valor:
            return "", "", "repaso"
        match = re.match(r"([BIA])(\d+)", valor)
        if match:
            nivel_map = {"B": "Básico", "I": "Intermedio", "A": "Avanzado"}
            return nivel_map.get(match.group(1), ""), match.group(2), None
        return "", "", None

    niveles = []
    ciclos = []
    overrides = []
    for valor in df.get("CICLO", []):
        try:
            result = extraer_nivel_y_ciclo(valor)
            if not isinstance(result, (list, tuple)) or len(result) != 3:
                result = ("", "", None)
        except Exception:
            result = ("", "", None)
        nivel, ciclo, override = result
        niveles.append(nivel)
        ciclos.append(ciclo)
        overrides.append(override)
    df["Nivel"] = niveles if niveles else None
    df["Ciclo"] = ciclos if ciclos else None
    df["_mod_override"] = overrides if overrides else None

    if "_mod_override" in df.columns and "MODALIDAD" in df.columns:
        df["MODALIDAD"] = df.apply(
            lambda row: row["_mod_override"] if pd.notna(row.get("_mod_override")) else row.get("MODALIDAD"),
            axis=1
        )
        df.drop(columns=["_mod_override"], inplace=True)

    # Días detectados
    def extraer_dias(texto):
        if pd.isna(texto): return []
        texto = texto.upper().replace(" Y ", ", ")
        dias_validos = ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES", "SÁBADOS", "DOMINGOS"]
        return [d for d in map(str.strip, texto.split(",")) if d in dias_validos]

    df["DÍAS DETECTADOS"] = df["DIAS"].apply(extraer_dias)

    # Inscritos y esperados
    def separar_inscritos(val):
        if pd.isna(val): return pd.Series([None, None])
        val = str(val).strip()
        if "/" in val:
            try:
                num, esperado = val.split("/")
                return pd.Series([int(num), int(esperado)])
            except:
                return pd.Series([None, None])
        elif val.isdigit():
            return pd.Series([int(val), None])
        return pd.Series([None, None])

    if "Nª inscritos" in df.columns:
        df[["N° Inscritos", "N° Esperado"]] = df["Nª inscritos"].apply(separar_inscritos)
    else:
        df["N° Inscritos"] = None
        df["N° Esperado"] = None

    # Horario detallado estructurado
    dia_a_codigo = {
        "LUNES": 0, "MARTES": 1, "MIÉRCOLES": 2,
        "JUEVES": 3, "VIERNES": 4, "SÁBADOS": 5, "DOMINGOS": 6
    }
    def parse_hora(hora_str):
        try:
            return datetime.strptime(hora_str.strip(), "%H:%M").time()
        except:
            return None

    def mapear_horarios_especial(dias, horas):
        if not isinstance(horas, str) or not dias:
            return {}
        bloques = [h.strip() for h in horas.split(",")]
        resultado = {}
        if len(bloques) == 2 and len(dias) >= 3:
            try:
                h1_inicio, h1_fin = map(parse_hora, bloques[0].split(" - "))
                h2_inicio, h2_fin = map(parse_hora, bloques[1].split(" - "))
                resultado[dia_a_codigo[dias[0]]] = (h1_inicio, h1_fin)
                for d in dias[1:]:
                    resultado[dia_a_codigo[d]] = (h2_inicio, h2_fin)
            except:
                return {}
        elif len(bloques) == 1:
            try:
                h_inicio, h_fin = map(parse_hora, bloques[0].split(" - "))
                for d in dias:
                    resultado[dia_a_codigo[d]] = (h_inicio, h_fin)
            except:
                return {}
        return resultado

    df["HORARIO DETALLADO"] = df.apply(lambda row: mapear_horarios_especial(row["DÍAS DETECTADOS"], row["HORAS"]), axis=1)

    # Fechas como date (con formato seguro)
    for col in ["F. Inicio", "F. Fin", "Parcial", "Final", "Subida de notas"]:
        df[col] = pd.to_datetime(df[col], format="%Y-%m-%d", errors='coerce').dt.date

    # Convertir columnas a enteros o nulo
    columnas_enteras = [
        "Ciclo", "N° Inscritos", "N° Esperado",
        "N° Aprobados", "N° Desaprobados", "N° No asistio (tiene 0)"
    ]
    for col in columnas_enteras:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")

    return df[COLUMNAS_FINALES].reset_index(drop=True)

def nombre_corto_curso(codigo_curso, df):
    fila = df[df["CODIGO"] == codigo_curso]
    if fila.empty:
        return f"❌ Código {codigo_curso} no encontrado."
    fila = fila.iloc[0]
    docente = str(fila["DOCENTE"]).strip()
    idioma = IDIOMA_ABBR.get(str(fila["IDIOMA"]).upper(), str(fila["IDIOMA"])[:3].upper())
    nivel = NIVEL_ABBR.get(fila["Nivel"], "NA")
    ciclo = str(fila["Ciclo"]).zfill(2) if pd.notna(fila["Ciclo"]) else "00"
    modalidad = MODALIDAD_ABBR.get(str(fila["MODALIDAD"]).lower(), "X")
    dias_abbr = "".join(sorted([DIA_COD.get(dia, "?") for dia in fila["HORARIO DETALLADO"].keys()]))
    horas_unicas = sorted(set([
        "-".join(f"{h.hour:02d}-{h.minute:02d}" for h in v if isinstance(h, time))
        for v in fila["HORARIO DETALLADO"].values()
        if isinstance(v, tuple) and all(isinstance(h, time) for h in v)
    ]))
    horario_final = "-".join(horas_unicas)
    return f"{docente}-{idioma} {modalidad}{ciclo}({nivel})-{dias_abbr}-{horario_final}"

def exportar_inscritos_formato_morado(
    codigo_curso,
    df_curso,
    feriados,
    plantilla_path="plantilla_lista_estudiantes.xlsx",
    carpeta_entrada="./",
    carpeta_salida="./"
):
    nombre_salida = nombre_corto_curso(codigo_curso, df_curso) + ".xlsx"
    nombre_salida = nombre_salida.replace("/", "-")
    output_dir = Path(carpeta_salida)
    output_dir.mkdir(parents=True, exist_ok=True)
    ruta_destino = str(output_dir / nombre_salida)

    shutil.copy(plantilla_path, ruta_destino)

    wb = load_workbook(ruta_destino)
    ws = wb.active
    df_inscritos = pd.read_excel(f"{carpeta_entrada}/Inscritos_{codigo_curso}.xlsx")

    n_estudiantes = df_inscritos.shape[0]

    # Copiar formato de la fila 2 (A-E) hacia abajo para cada estudiante
    for i in range(n_estudiantes):
        source_row = 2
        target_row = 2 + i
        for col in range(1, 6):  # columnas A-E
            cell_src = ws.cell(row=source_row, column=col)
            cell_tgt = ws.cell(row=target_row, column=col)
            cell_tgt._style = copy(cell_src._style)
            cell_tgt.font = copy(cell_src.font)
            cell_tgt.border = copy(cell_src.border)
            cell_tgt.fill = copy(cell_src.fill)
            cell_tgt.number_format = copy(cell_src.number_format)
            cell_tgt.protection = copy(cell_src.protection)
            cell_tgt.alignment = copy(cell_src.alignment)

    # Llenar datos en las filas A3-E{n}
    for idx, row in enumerate(df_inscritos.itertuples(index=False), start=2):
        ws[f"A{idx}"] = idx - 1
        ws[f"B{idx}"] = row.CODIGO_CURSO
        ws[f"C{idx}"] = row.NOMBRES
        ws[f"D{idx}"] = row.CORREO
        ws[f"E{idx}"] = row.CELULAR

    # Poner modalidad, nivel, ciclo en G2, fechas en H2 e I2
    fila_curso = df_curso[df_curso["CODIGO"] == codigo_curso].iloc[0]
    nivel = fila_curso["Nivel"]
    ciclo = str(fila_curso["Ciclo"]).zfill(2) if pd.notna(fila_curso["Ciclo"]) else ""
    modalidad_nivel_ciclo = f"{MODALIDAD_ABBR.get(str(fila_curso['MODALIDAD']).lower(), 'X')} {NIVEL_ABBR.get(nivel, 'X')}{ciclo}"
    ws["G2"] = modalidad_nivel_ciclo
    ws["H2"] = pd.to_datetime(fila_curso["F. Inicio"])
    ws["I2"] = pd.to_datetime(fila_curso["F. Fin"])
    ws["H2"].number_format = 'DD-MMM'
    ws["I2"].number_format = 'DD-MMM'

    # Poner feriados debajo de "Feriados" en G4 para abajo
    ws["G4"] = "Feriados"
    for i, f in enumerate(feriados):
        ws.cell(row=5+i, column=7).value = f

    wb.save(ruta_destino)
    print("✅ Exportado:", ruta_destino)
    return ruta_destino

def crear_carpeta_salida():
    output_dir = Path("./output")
    if not output_dir.exists():
        output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir

def crear_carpeta_inscritos():
    inscritos_dir = Path("./inscritos")
    if not inscritos_dir.exists():
        inscritos_dir.mkdir(parents=True, exist_ok=True)
    return inscritos_dir

# ========== INTERFAZ DE CONFIGURACIÓN ==========

def seleccionar_plantilla(config):
    limpiar_consola()
    archivos = [f for f in Path('.').glob("*.xlsx") if not str(f).startswith("~$")]
    if not archivos:
        print("❌ No se encontraron archivos .xlsx en la carpeta actual.")
        pausar()
        return None
    opciones = [{"name": f.name, "value": str(f)} for f in archivos]
    plantilla = inquirer.select(
        message="Selecciona la plantilla (.xlsx):",
        choices=opciones,
    ).execute()
    config["plantilla"] = plantilla
    guardar_config(config)
    return plantilla

def seleccionar_carga_horaria(config):
    limpiar_consola()
    archivos = [f for f in Path('.').glob("*.xlsx") if not str(f).startswith("~$")]
    if not archivos:
        print("❌ No se encontraron archivos .xlsx en la carpeta actual.")
        pausar()
        return None
    opciones = [{"name": f.name, "value": str(f)} for f in archivos]
    carga_horaria = inquirer.select(
        message="Selecciona el archivo de carga horaria (.xlsx):",
        choices=opciones,
    ).execute()
    config["carga_horaria"] = carga_horaria
    guardar_config(config)
    return carga_horaria

def seleccionar_mes(config, carga_horaria):
    limpiar_consola()
    xl = pd.ExcelFile(carga_horaria)
    opciones = [{"name": sh, "value": sh} for sh in xl.sheet_names]
    mes = inquirer.select(
        message="Selecciona el mes (sheet):",
        choices=opciones,
    ).execute()
    config["mes"] = mes
    guardar_config(config)
    return mes

def ingresar_feriados(config):
    limpiar_consola()
    print("Ejemplo de feriados: 2025-06-29,2025-07-28")
    feriados = inquirer.text(message="Feriados a incluir (separados por coma, ENTER si ninguno): ").execute()
    feriados = [x.strip() for x in feriados.split(",") if x.strip()]
    config["feriados"] = feriados
    guardar_config(config)
    return feriados

def mostrar_config(config):
    print("\n======= Configuración Actual =======")
    for k, v in config.items():
        print(f"{k}: {v}")

def seleccionar_varios_archivos(archivos_validos, df_cursos):
    if not archivos_validos:
        print("❌ No se encontraron archivos de inscritos coincidentes con los cursos.")
        pausar()
        return []
    opciones = [
        {"name": f"{nombre_corto_curso(int(cod), df_cursos)} ({f.name})", "value": (cod, f)}
        for cod, f in archivos_validos
    ]
    opciones.append(Separator())
    opciones.append({"name": "Seleccionar todos", "value": "all"})
    seleccionados = inquirer.checkbox(
        message="Selecciona los archivos a exportar (Espacio para marcar/desmarcar, ENTER para confirmar):",
        choices=opciones,
        instruction="Usa ↑/↓ para navegar, espacio para marcar"
    ).execute()
    if "all" in seleccionados:
        return archivos_validos
    return seleccionados

# ========== MAIN CON MENÚ ==========
def main():
    crear_carpeta_salida()
    crear_carpeta_inscritos()
    config = cargar_config()
    while True:
        limpiar_consola()
        op = inquirer.select(
            message="===== Exportador de Inscritos con Formato Morado =====",
            choices=[
                {"name": "Exportar cursos", "value": "1"},
                {"name": "Cambiar archivo de carga horaria", "value": "2"},
                {"name": "Cambiar mes/sheet", "value": "3"},
                {"name": "Ingresar feriados", "value": "4"},
                {"name": "Cambiar plantilla", "value": "5"},
                {"name": "Mostrar carga horaria y cursos", "value": "6"},
                {"name": "Mostrar configuración actual", "value": "7"},
                {"name": "Salir", "value": "0"},
            ],
            default="1",
        ).execute()


        if op == "0":
            break
        elif op == "1":
            # Requisitos mínimos
            if not all(k in config for k in ["plantilla", "carga_horaria", "mes"]):
                print("⚠️ Configura primero plantilla, archivo de carga horaria y mes.")
                pausar()
                continue
            limpiar_consola()
            print("Leyendo cursos...")
            df_cursos = clean_df_mes_idioma(config["carga_horaria"], config["mes"])
            if df_cursos.empty:
                print("❌ No se encontraron cursos para el mes/archivo seleccionado.")
                pausar()
                continue

            limpiar_consola()
            print("Buscando archivos en inscritos/")
            inscritos_folder = Path("inscritos/")
            if not inscritos_folder.exists():
                print("❌ La carpeta inscritos/ no existe.")
                pausar()
                continue
            archivos_inscritos = list(inscritos_folder.glob("Inscritos_*.xlsx"))
            print("Archivos encontrados:", ", ".join([f.name for f in archivos_inscritos]) if archivos_inscritos else "Ninguno")

            codigos_cursos_str = set(str(c) for c in df_cursos["CODIGO"])
            codigos_cursos_int = set()
            try:
                codigos_cursos_int = set(int(c) for c in df_cursos["CODIGO"])
            except Exception:
                pass

            archivos_validos = []
            for f in archivos_inscritos:
                cod_match = f.stem.split("_")[-1]
                if (cod_match in codigos_cursos_str) or (cod_match.isdigit() and int(cod_match) in codigos_cursos_int):
                    archivos_validos.append((cod_match, f))

            seleccionados = seleccionar_varios_archivos(archivos_validos, df_cursos)
            if not seleccionados:
                continue

            feriados = config.get("feriados", [])
            print("Exportando los siguientes cursos:")
            for cod, f in seleccionados:
                desc = nombre_corto_curso(int(cod), df_cursos)
                print(f"- {desc}")

            for cod, f in seleccionados:
                try:
                    exportar_inscritos_formato_morado(
                        int(cod), df_cursos, feriados,
                        plantilla_path=config["plantilla"],
                        carpeta_entrada="inscritos/",
                        carpeta_salida="./output/"
                    )
                except Exception as e:
                    print(f"❌ Error exportando {f.name}: {e}")
            pausar()
        elif op == "2":
            seleccionar_carga_horaria(config)
        elif op == "3":
            if "carga_horaria" not in config:
                print("Primero selecciona archivo de carga horaria.")
                pausar()
            else:
                seleccionar_mes(config, config["carga_horaria"])
        elif op == "4":
            ingresar_feriados(config)
        elif op == "5":
            seleccionar_plantilla(config)
        elif op == "6":
            if not all(k in config for k in ["plantilla", "carga_horaria", "mes"]):
                print("⚠️ Configura primero plantilla, archivo de carga horaria y mes.")
                pausar()
                continue
            limpiar_consola()
            print("Leyendo cursos...")
            df_cursos = clean_df_mes_idioma(config["carga_horaria"], config["mes"])
            if df_cursos.empty:
                print("❌ No se encontraron cursos para el mes/archivo seleccionado.")
                pausar()
                continue
            print("Cursos encontrados:\n")
            print(df_cursos[["CODIGO", "Nivel", "Ciclo", "MODALIDAD", "DOCENTE", "IDIOMA", "F. Inicio", "F. Fin", "N° Inscritos"]].to_string(index=False))
            pausar()
        elif op == "7":
            mostrar_config(config)
            pausar()

if __name__ == "__main__":
    main()
